import shutil
import tempfile
from pathlib import Path
import openpyxl
import pandas as pd
import pytest
from engine.excel_writer import generate_excel
from engine.models import AnalysisResult, Assumptions, Multiples, Scenarios
from engine.calculator import build_year_col_map

TEMPLATE = Path("templates/Modelo 2025 vacio.xlsx")
DATES = pd.to_datetime(["2022-09-30", "2023-09-30", "2024-09-30"])

def make_result():
    income = pd.DataFrame({
        DATES[2]: {"Total Revenue": 391000000000, "Operating Income": 123000000000,
                   "Net Income": 94000000000, "Tax Provision": 30000000000,
                   "Diluted Average Shares": 15000000000,
                   "Interest Expense": -3000000000},
    }).T
    balance = pd.DataFrame({
        DATES[2]: {"Cash And Cash Equivalents": 65000000000, "Total Debt": 98000000000,
                   "Stockholders Equity": 56000000000, "Goodwill": 0},
    }).T
    cashflow = pd.DataFrame({
        DATES[2]: {"Depreciation And Amortization": 11700000000,
                   "Operating Cash Flow": 118000000000,
                   "Capital Expenditure": -9500000000},
    }).T
    return AnalysisResult(
        ticker="TEST",
        company_name="Test Corp",
        sector="Technology",
        industry="Software",
        currency="USD",
        years=list(DATES),
        income=income,
        balance=balance,
        cashflow=cashflow,
        history=pd.DataFrame(),
        assumptions=Assumptions(growth=0.08, ebit_margin=0.31, tax_rate=0.18,
                                price=293.0, div_yield=0.004, buyback_pct=0.87),
        multiples=Multiples(per=30.0, pfcf=28.0, ev_ebitda=24.0, ev_ebit=26.0),
        scenarios=Scenarios(bull_3y=400.0, bull_5y=550.0, mid_3y=320.0,
                            mid_5y=400.0, bear_3y=200.0, bear_5y=220.0),
        year_col=build_year_col_map(list(DATES)),
    )

def test_generate_excel_creates_file():
    result = make_result()
    with tempfile.TemporaryDirectory() as tmp:
        output = Path(tmp) / "TEST_Modelo_2025.xlsx"
        generate_excel(result, TEMPLATE, output)
        assert output.exists()

def test_generate_excel_fills_price():
    result = make_result()
    with tempfile.TemporaryDirectory() as tmp:
        output = Path(tmp) / "TEST_Modelo_2025.xlsx"
        generate_excel(result, TEMPLATE, output)
        wb = openpyxl.load_workbook(output)
        ws_val = wb[wb.sheetnames[4]]
        assert ws_val["Q9"].value == 293.0

def test_generate_excel_fills_multiples():
    result = make_result()
    with tempfile.TemporaryDirectory() as tmp:
        output = Path(tmp) / "TEST_Modelo_2025.xlsx"
        generate_excel(result, TEMPLATE, output)
        wb = openpyxl.load_workbook(output)
        ws_val = wb[wb.sheetnames[4]]
        assert ws_val["Q21"].value == 30.0
        assert ws_val["Q22"].value == 28.0
