import shutil
from pathlib import Path
import openpyxl
from engine.models import AnalysisResult
from engine.calculator import (
    get_value, DA_KEYS, CASH_KEYS, EBIT_KEYS, DEBT_KEYS,
    OPCF_KEYS, CAPEX_KEYS, BUY_KEYS, SHARE_KEYS, FIRST_COL, LAST_COL
)

INCOME_MAP = {
     5: (["Total Revenue", "TotalRevenue", "Revenue"],                          1),
     6: (["Cost Of Revenue", "CostOfRevenue"],                                 -1),
     8: (["Gross Profit", "GrossProfit"],                                       1),
    12: (EBIT_KEYS,                                                             1),
    14: (["Interest Expense", "InterestExpense"],                              -1),
    22: (["Tax Provision", "IncomeTaxExpense"],                                -1),
    24: (["Net Income", "NetIncome"],                                           1),
    31: (EBIT_KEYS,                                                             1),
}
BALANCE_MAP = {
    39: (CASH_KEYS,                                                             1),
    48: (["Current Assets", "Total Current Assets"],                            1),
    49: (["Net PPE", "NetPPE"],                                                 1),
    52: (["Goodwill"],                                                          1),
    59: (["Total Assets", "TotalAssets"],                                       1),
    68: (["Current Liabilities", "Total Current Liabilities"],                  1),
    80: (["Stockholders Equity", "Common Stock Equity"],                        1),
    83: (DEBT_KEYS,                                                             1),
}
CF_MAP = {
     98: (DA_KEYS,                                                              1),
    115: (OPCF_KEYS,                                                            1),
    116: (CAPEX_KEYS,                                                           1),
    124: (BUY_KEYS,                                                             1),
}


def _set(ws, row, col, value):
    if value is not None:
        try:
            ws.cell(row=row, column=col, value=float(value))
        except (TypeError, ValueError):
            ws.cell(row=row, column=col, value=value)


def _clear_inputs(wb):
    val_sheet = wb.sheetnames[4]
    ws_is = wb["1.Income statement"]
    for c in ("P11", "P16", "P21"):
        ws_is[c] = None

    ws_fc = wb["2.Flujos de caja"]
    for col in ("J", "K", "L", "M", "N"):
        ws_fc[f"{col}24"] = f"='{val_sheet}'!Q9"
        ws_fc[f"{col}26"] = None

    ws_val = wb[val_sheet]
    for c in ("Q9", "Q11", "Q21", "Q22", "Q23", "Q24"):
        ws_val[c] = None
    for r in (29, 30, 31):
        for c in ("M", "N", "P"):
            ws_val[f"{c}{r}"] = None


def _fill_datos_tweenvest(ws, result: AnalysisResult):
    for r in range(2, 149):
        for c in range(FIRST_COL, LAST_COL + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and not str(cell.value).startswith("="):
                cell.value = None

    for date, col in result.year_col.items():
        ws.cell(row=2, column=col, value=f"{date.year} FY")
        ws.cell(row=3, column=col, value=date.date())

    for row_idx, (keys, sign) in INCOME_MAP.items():
        for date, col in result.year_col.items():
            v = get_value(result.income, keys, date)
            if v is not None:
                _set(ws, row_idx, col, v * sign)

    for date, col in result.year_col.items():
        ebit = get_value(result.income, EBIT_KEYS, date)
        da   = get_value(result.cashflow, DA_KEYS, date)
        if ebit is not None and da is not None:
            _set(ws, 29, col, ebit + abs(da))

    for row_idx, (keys, sign) in BALANCE_MAP.items():
        for date, col in result.year_col.items():
            v = get_value(result.balance, keys, date)
            if v is not None:
                _set(ws, row_idx, col, v * sign)

    for date, col in result.year_col.items():
        debt = get_value(result.balance, DEBT_KEYS, date)
        cash = get_value(result.balance, CASH_KEYS, date)
        if debt is not None and cash is not None:
            _set(ws, 84, col, debt - cash)

    for row_idx, (keys, sign) in CF_MAP.items():
        for date, col in result.year_col.items():
            v = get_value(result.cashflow, keys, date)
            if v is not None:
                _set(ws, row_idx, col, v * sign)

    for date, col in result.year_col.items():
        v = get_value(result.income, SHARE_KEYS, date)
        if v is not None:
            _set(ws, 143, col, v)


def _fill_assumptions(wb, result: AnalysisResult):
    a = result.assumptions
    ws_is = wb["1.Income statement"]
    ws_is["P11"] = round(a.growth, 4)
    ws_is["P16"] = round(a.ebit_margin, 4)
    ws_is["P21"] = round(a.tax_rate, 4)

    val_sheet = wb.sheetnames[4]
    ws_val = wb[val_sheet]
    ws_val["Q9"]  = round(a.price, 2) if a.price else None
    ws_val["Q11"] = round(a.div_yield, 4) if a.div_yield else None

    m = result.multiples
    if m.per:       ws_val["Q21"] = m.per
    if m.pfcf:      ws_val["Q22"] = m.pfcf
    if m.ev_ebitda: ws_val["Q23"] = m.ev_ebitda
    if m.ev_ebit:   ws_val["Q24"] = m.ev_ebit

    s = result.scenarios
    ws_val["M29"] = s.bull_3y; ws_val["N29"] = s.bull_5y
    ws_val["M30"] = s.mid_3y;  ws_val["N30"] = s.mid_5y
    ws_val["M31"] = s.bear_3y; ws_val["N31"] = s.bear_5y
    ws_val["P29"] = 0.20
    ws_val["P30"] = 0.60
    ws_val["P31"] = 0.20

    ws_fc = wb["2.Flujos de caja"]
    for col in ("J", "K", "L", "M", "N"):
        ws_fc[f"{col}26"] = a.buyback_pct


def generate_excel(result: AnalysisResult, template: Path, output: Path) -> Path:
    """Genera Excel rellenado a partir de AnalysisResult. Devuelve path del fichero."""
    shutil.copy(template, output)
    wb = openpyxl.load_workbook(output)
    _clear_inputs(wb)
    _fill_datos_tweenvest(wb["Datos Tweenvest"], result)
    _fill_assumptions(wb, result)
    wb.save(output)
    return output
